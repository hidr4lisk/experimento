from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_http_methods
from .models import Agent, Record
from datetime import datetime, timedelta, date
import holidays
from django.core.management import call_command
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

@login_required(login_url='login')
def export_agent_report(request, agent_id):
    agent = get_object_or_404(Agent, id=agent_id)
    records = Record.objects.filter(agent=agent).order_by('fecha_inicio')
    
    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Reporte {agent.name[:20]}"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    
    thin_side = Side(style='thin', color="000000")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # Encabezados
    headers = ['Agente', 'Tipo de Licencia', 'Fecha Inicio', 'Fecha Fin', 'Notas']
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # Datos
    for row_num, record in enumerate(records, 2):
        cells = [
            ws.cell(row=row_num, column=1, value=agent.name),
            ws.cell(row=row_num, column=2, value=record.get_record_type_display()),
            ws.cell(row=row_num, column=3, value=record.fecha_inicio.strftime('%d/%m/%Y')),
            ws.cell(row=row_num, column=4, value=record.fecha_fin.strftime('%d/%m/%Y')),
            ws.cell(row=row_num, column=5, value=record.notes or "-")
        ]
        for cell in cells:
            cell.alignment = center_align
            cell.border = thin_border
    
    # Ajustar ancho de columnas (Tabla principal)
    for i in range(1, 6):
        max_length = 0
        column = get_column_letter(i)
        for cell in ws[column]:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 5

    # Preparar respuesta
    filename = "Reporte SIA.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

def calculate_agent_status(agent):
    """
    Calcula el estado de disponibilidad de un agente.
    
    Retorna un diccionario con:
    - 'available': True si el agente está disponible, False si está de licencia
    - 'return_date': fecha de reintegro (solo si available=False), calculada
                     considerando solo días hábiles (lun-vie, excluyendo feriados)
    """
    today = date.today()
    
    # Buscar SOLO registros activos que incluyan HOY (no futuros)
    # Un agente está de licencia solo si tiene un registro activo en este momento
    active_records = Record.objects.filter(
        agent=agent,
        fecha_inicio__lte=today,
        fecha_fin__gte=today
    ).order_by('fecha_inicio')
    
    if not active_records.exists():
        # No hay registros activos hoy = está disponible
        # (aunque tenga vacaciones programadas para el futuro)
        return {'available': True, 'return_date': None}
    
    # Encontrar el último día de licencia de los registros activos
    last_end_date = max(r.fecha_fin for r in active_records)
    
    # Calcular siguiente día hábil después del último día de licencia
    return_date = last_end_date + timedelta(days=1)
    ar_holidays = holidays.AR(years=range(return_date.year, return_date.year + 2))
    
    # Saltar fines de semana y feriados
    while return_date.weekday() >= 5 or return_date in ar_holidays:
        return_date += timedelta(days=1)
        # Actualizar feriados si cambiamos de año
        if return_date.year not in ar_holidays.years:
            ar_holidays = holidays.AR(years=range(return_date.year, return_date.year + 2))
    
    return {'available': False, 'return_date': return_date}

def home(request):
    if not request.user.is_authenticated:
        return render(request, 'experimentapp/home_not_logged.html')
    
    records = Record.objects.all().order_by('-fecha_inicio')
    
    # Vista actual (asistencia o agentes)
    current_view = request.GET.get('view', 'asistencia')
    
    # Filtros
    agent_filter = request.GET.get('agent')
    record_type_filter = request.GET.get('type')
    search_query = request.GET.get('search')
    show_duplicates = request.GET.get('duplicates') == 'true'
    
    if agent_filter and agent_filter != 'all':
        records = records.filter(agent_id=agent_filter)
    
    if record_type_filter:
        records = records.filter(record_type=record_type_filter)
    if search_query:
        from django.db.models import Q
        records = records.filter(
            Q(agent__name__icontains=search_query) | 
            Q(agent__location__icontains=search_query) | 
            Q(notes__icontains=search_query)
        )
    
    if show_duplicates:
        from django.db.models import Exists, OuterRef
        overlapping = Record.objects.filter(
            agent=OuterRef('agent'),
            fecha_inicio__lte=OuterRef('fecha_fin'),
            fecha_fin__gte=OuterRef('fecha_inicio')
        ).exclude(id=OuterRef('id'))
        
        records = records.annotate(
            has_overlap=Exists(overlapping)
        ).filter(has_overlap=True)
    
    # Ordenamiento
    sort_by = request.GET.get('sort', 'name' if current_view == 'agentes' else '-fecha_inicio')
    
    # Check if any filter is applied (including the 'all' selection for agents)
    has_filter = any([
        agent_filter is not None, 
        record_type_filter, 
        search_query, 
        show_duplicates
    ])
    
    if current_view == 'asistencia' and not has_filter:
        records = Record.objects.none()
    elif current_view == 'asistencia':
        records = records.order_by(sort_by)

    agents = Agent.objects.all().order_by('name')
    if current_view == 'agentes':
        agents = agents.order_by(sort_by)
    
    is_admin = request.user.is_superuser
    is_editor = request.user.is_superuser or request.user.is_staff
    
    # Lógica para "De Licencia"
    licensing_agents = []
    if current_view == 'de_licencia':
        for agent in agents:
            status = calculate_agent_status(agent)
            if not status['available']:
                licensing_agents.append({
                    'agent': agent,
                    'return_date': status['return_date']
                })

    context = {
        'records': records,
        'agents': agents,
        'licensing_agents': licensing_agents,
        'is_admin': is_admin,
        'is_editor': is_editor,
        'record_types': Record.RECORD_TYPES,
        'current_view': current_view,
        'has_filter': has_filter,
    }
    return render(request, 'experimentapp/home.html', context)

@login_required(login_url='login')
def agent_calendar(request, agent_id):
    agent = get_object_or_404(Agent, id=agent_id)
    records = Record.objects.filter(agent=agent)
    
    # Obtener todos los agentes para el dropdown
    all_agents = Agent.objects.all().order_by('name')
    
    # Calcular estado de disponibilidad del agente
    agent_status = calculate_agent_status(agent)
    
    context = {
        'agent': agent,
        'records': records,
        'all_agents': all_agents,
        'agent_status': agent_status,
    }
    return render(request, 'experimentapp/agent_calendar.html', context)

@login_required(login_url='login')
def agent_calendar_data(request, agent_id):
    """API que devuelve los eventos del calendario en formato JSON"""
    agent = get_object_or_404(Agent, id=agent_id)
    records = Record.objects.filter(agent=agent)
    
    # Mapeo de colores por tipo de registro
    color_map = {
        'vacaciones': '#ffc107',      # Amarillo
        'franquicia': '#6f42c1',      # Morado
        'razon_particular': '#fd7e14', # Naranja
        'comision': '#dc3545',        # Rojo
    }
    
    events = []
    
    # Obtener el rango de fechas para los feriados
    all_dates = [r.fecha_inicio for r in records] + [r.fecha_fin for r in records]
    if all_dates:
        min_year = min(d.year for d in all_dates)
        max_year = max(d.year for d in all_dates)
    else:
        min_year = datetime.now().year
        max_year = datetime.now().year
    
    # Cargar feriados de Argentina para los años relevantes
    ar_holidays = holidays.AR(years=range(min_year, max_year + 1))
    
    # Agregar feriados al calendario
    for date, name in ar_holidays.items():
        if date.weekday() < 5:  # Solo feriados en días hábiles (opcional, pero consistente con el resto)
            events.append({
                'title': f'Feriado: {name}',
                'start': date.strftime('%Y-%m-%d'),
                'backgroundColor': '#e2e8f0',
                'borderColor': '#cbd5e0',
                'textColor': '#4a5568',
                'allDay': True,
                'display': 'background', # Opcional: mostrar como fondo
            })
        else:
            # Si es fin de semana, igual lo agregamos pero quizás con otro estilo o solo el título
            events.append({
                'title': f'Feriado: {name}',
                'start': date.strftime('%Y-%m-%d'),
                'backgroundColor': '#e2e8f0',
                'borderColor': '#cbd5e0',
                'textColor': '#4a5568',
                'allDay': True,
            })

    # Agregar eventos de los registros (días que no vino)
    for record in records:
        current_date = record.fecha_inicio
        color = color_map.get(record.record_type, '#667eea')  # Color default
        
        while current_date <= record.fecha_fin:
            # Solo marcamos días de lunes a viernes (weekday 0-4)
            if current_date.weekday() < 5:  # 0=lunes, 4=viernes
                events.append({
                    'title': f'{record.get_record_type_display()}',
                    'start': current_date.strftime('%Y-%m-%d'),
                    'backgroundColor': color,
                    'borderColor': color,
                    'textColor': 'white',
                    'allDay': True,
                })
            current_date += timedelta(days=1)
    
    return JsonResponse(events, safe=False)

def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect('home')
        context = {'error': 'Credenciales inválidas'}
        return render(request, "experimentapp/login.html", context)
    return render(request, "experimentapp/login.html")

@login_required(login_url='login')
def logout_view(request):
    logout(request)
    return redirect('login')

@login_required(login_url='login')
def add_agent(request):
    if not (request.user.is_superuser or request.user.is_staff):
        return HttpResponse("No autorizado", status=403)
    
    if request.method == "POST":
        name = request.POST.get('name')
        location = request.POST.get('location')
        Agent.objects.create(name=name, location=location)
        return redirect('home')
    
    return render(request, 'experimentapp/add_agent.html')

@login_required(login_url='login')
def edit_agent(request, agent_id):
    if not (request.user.is_superuser or request.user.is_staff):
        return HttpResponse("No autorizado", status=403)
    
    agent = get_object_or_404(Agent, id=agent_id)
    
    if request.method == "POST":
        agent.name = request.POST.get('name')
        agent.location = request.POST.get('location')
        agent.save()
        return redirect('home')
    
    context = {'agent': agent}
    return render(request, 'experimentapp/edit_agent.html', context)

@login_required(login_url='login')
def add_record(request):
    if not (request.user.is_superuser or request.user.is_staff):
        return HttpResponse("No autorizado", status=403)
    
    if request.method == "POST":
        agent_id = request.POST.get('agent')
        record_type = request.POST.get('record_type')
        notes = request.POST.get('notes', '')
        
        fechas_inicio = request.POST.getlist('fecha_inicio[]')
        fechas_fin = request.POST.getlist('fecha_fin[]')
        
        # Fallback si el JS no se usó y enviaron "fecha_inicio" normal
        if not fechas_inicio:
            fechas_inicio = request.POST.getlist('fecha_inicio')
        if not fechas_fin:
            fechas_fin = request.POST.getlist('fecha_fin')

        def parse_date(date_str):
            if not date_str: return None
            try:
                return datetime.strptime(date_str, '%d/%m/%Y').strftime('%Y-%m-%d')
            except ValueError:
                return date_str

        agent = get_object_or_404(Agent, id=agent_id)
        
        for i in range(len(fechas_inicio)):
            f_inicio = parse_date(fechas_inicio[i])
            if not f_inicio: continue
            
            f_fin_str = fechas_fin[i] if i < len(fechas_fin) else None
            f_fin = parse_date(f_fin_str) if f_fin_str else f_inicio
            
            Record.objects.create(
                agent=agent,
                record_type=record_type,
                fecha_inicio=f_inicio,
                fecha_fin=f_fin,
                notes=notes
            )
        return redirect('home')
    
    agents = Agent.objects.all().order_by('name')
    context = {
        'agents': agents,
        'record_types': Record.RECORD_TYPES
    }
    return render(request, 'experimentapp/add_record.html', context)

@login_required(login_url='login')
def edit_record(request, record_id):
    if not (request.user.is_superuser or request.user.is_staff):
        return HttpResponse("No autorizado", status=403)
    
    record = get_object_or_404(Record, id=record_id)
    
    if request.method == "POST":
        def parse_date(date_str):
            if not date_str: return None
            try:
                return datetime.strptime(date_str, '%d/%m/%Y').strftime('%Y-%m-%d')
            except ValueError:
                return date_str

        record.agent_id = request.POST.get('agent')
        record.record_type = request.POST.get('record_type')
        f_inicio = parse_date(request.POST.get('fecha_inicio'))
        f_fin = parse_date(request.POST.get('fecha_fin'))
        
        record.fecha_inicio = f_inicio
        record.fecha_fin = f_fin if f_fin else f_inicio
        record.notes = request.POST.get('notes', '')
        record.save()
        return redirect('home')
    
    agents = Agent.objects.all().order_by('name')
    context = {
        'record': record,
        'agents': agents,
        'record_types': Record.RECORD_TYPES
    }
    return render(request, 'experimentapp/edit_record.html', context)  

@login_required(login_url='login')
def delete_record(request, record_id):
    if not (request.user.is_superuser or request.user.is_staff):
        return HttpResponse("No autorizado", status=403)
    
    record = get_object_or_404(Record, id=record_id)
    record.delete()
    return redirect('home')

@login_required(login_url='login')
def delete_agent(request, agent_id):
    if not (request.user.is_superuser or request.user.is_staff):
        return HttpResponse("No autorizado", status=403)
    
    agent = get_object_or_404(Agent, id=agent_id)
    agent.delete()
    return redirect('home')

def debug_db(request):
    """View para diagnosticar problemas de base de datos"""
    # No verificamos autenticación aquí para poder acceder aun cuando el login falla
    # Pero usamos un secreto o algo simple si es necesario, por ahora acceso libre
    # para debug rápido (solo habilitar temporalmente)
    try:
        from django.db import connection
        from django.core.management import call_command
        
        # Si pasan ?migrate=true, forzamos las migraciones
        if request.GET.get('migrate') == 'true':
            if not request.user.is_superuser:
                return JsonResponse({'error': 'No autorizado'}, status=403)
            call_command('migrate', no_input=True)
            return JsonResponse({
                'status': 'Migrations attempted',
                'migration_output': 'OK',
            })

        with connection.cursor() as cursor:
            cursor.execute("SELECT 1")
            row = cursor.fetchone()
        
        from .models import Agent, Record
        from django.template.loader import render_to_string
        
        # Test rendering a template
        try:
            rendered_login = render_to_string("experimentapp/login.html")
            render_test = f"OK (Size: {len(rendered_login)})"
        except Exception as te:
            render_test = f"ERROR: {str(te)}"
        
        db_info = {
            'connection': 'OK' if row else 'Failed',
            'engine': connection.settings_dict['ENGINE'],
            'host': connection.settings_dict['HOST'],
            'user_count': User.objects.count(),
            'agent_count': Agent.objects.count(),
            'record_count': Record.objects.count(),
            'render_test_login': render_test,
        }
        return JsonResponse(db_info)
    except Exception as e:
        import traceback
        return JsonResponse({
            'error': str(e),
            'traceback': traceback.format_exc()
        }, status=500)
@login_required(login_url='login')
def trigger_population(request):
    if not (request.user.is_superuser or request.user.is_staff):
        return HttpResponse("No autorizado", status=403)
    
    try:
        from django.core.management import call_command
        call_command('populate_agents')
        return HttpResponse("¡Éxito! La base de datos ha sido poblada con la lista definitiva.")
    except Exception as e:
        return HttpResponse(f"Error al poblar la base de datos: {str(e)}", status=500)

@login_required(login_url='login')
def manage_users(request):
    if not request.user.is_superuser:
        return HttpResponse("No autorizado", status=403)
    
    users = User.objects.all().order_by('username')
    return render(request, 'experimentapp/manage_users.html', {'users': users})

@login_required(login_url='login')
def edit_user_role(request, user_id):
    if not request.user.is_superuser:
        return HttpResponse("No autorizado", status=403)
    
    user = get_object_or_404(User, id=user_id)
    if request.method == "POST":
        role = request.POST.get('role')
        if role == 'editor':
            user.is_superuser = False
            user.is_staff = True
        else: # lector
            user.is_superuser = False
            user.is_staff = False
        user.save()
        return redirect('manage_users')
    
    return render(request, 'experimentapp/edit_user.html', {'target_user': user})

@login_required(login_url='login')
def change_password(request, user_id):
    if not request.user.is_superuser:
        return HttpResponse("No autorizado", status=403)
    
    user = get_object_or_404(User, id=user_id)
    if request.method == "POST":
        new_password = request.POST.get('password')
        if new_password:
            user.set_password(new_password)
            user.save()
            return redirect('manage_users')
    
    return render(request, 'experimentapp/edit_user.html', {
        'target_user': user,
        'mode': 'password'
    })

@login_required(login_url='login')
def delete_user(request, user_id):
    if not request.user.is_superuser:
        return HttpResponse("No autorizado", status=403)
    
    user = get_object_or_404(User, id=user_id)
    if user.is_superuser and User.objects.filter(is_superuser=True).count() <= 1:
        return HttpResponse("No puedes eliminar al único superusuario", status=400)
    
    user.delete()
    return redirect('manage_users')

@login_required(login_url='login')
def add_user(request):
    if not request.user.is_superuser:
        return HttpResponse("No autorizado", status=403)
    
    if request.method == "POST":
        username = request.POST.get('username')
        password = request.POST.get('password')
        role = request.POST.get('role')
        
        if User.objects.filter(username=username).exists():
            return HttpResponse("El usuario ya existe", status=400)
            
        user = User.objects.create_user(username=username, password=password)
        if role == 'editor':
            user.is_staff = True
        user.save()
        return redirect('manage_users')
    
    return render(request, 'experimentapp/add_user.html')
